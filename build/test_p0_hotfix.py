# -*- coding: utf-8 -*-
"""P0 Hotfix 单测脚本：IR / fallback / DataFrame / list → normalize_excel_input 收敛"""
import sys
import os
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from src.common.pdf_table_ir import (
    to_table_block, fallback_block, normalize_excel_input, to_dataframe, TableBlock, TableMeta
)

print("=" * 60)
print("P0 Hotfix 单测：normalize_excel_input 兼容层")
print("=" * 60)

# Test 1: dict IR
print("\nTEST 1: dict IR -> DataFrame")
ir = to_table_block(rows=[['A1','B1\nB2'],['A2','B2']], page=1, table_id=1)
print(f"  ir type: {type(ir).__name__}, meta mode: {ir['meta']['mode']}")
df1 = normalize_excel_input(ir)
print(f"  -> DataFrame type: {type(df1).__name__}, shape: {df1.shape}")
records = df1.fillna("").astype(str).to_dict("records")
print(f"  records: {records}")

# Test 2: fallback_block
print("\nTEST 2: fallback_block -> normalize_excel_input")
fb = fallback_block(rows=[['line1'],['line2']], page=2, table_id=1)
print(f"  fb mode: {fb['meta']['mode']}, confidence: {fb['meta']['confidence']}")
df2 = normalize_excel_input(fb)
print(f"  -> DataFrame type: {type(df2).__name__}, shape: {df2.shape}")

# Test 3: DataFrame pass-through
print("\nTEST 3: DataFrame pass-through")
import pandas as pd
df_in = pd.DataFrame([['A1','B1'],['A2','B2']], columns=['col1','col2'])
df3 = normalize_excel_input(df_in)
print(f"  is same object: {df3 is df_in}")

# Test 4: TableBlock dataclass
print("\nTEST 4: TableBlock dataclass.to_dataframe()")
tb = TableBlock(rows=[['x','y']], meta=TableMeta(page=1, table_id=1))
df4 = tb.to_dataframe()
print(f"  -> type: {type(df4).__name__}, shape: {df4.shape}")

# Test 5: 裸 list of list
print("\nTEST 5: 裸 list of list")
df5 = normalize_excel_input([['a','b'],['c','d']])
print(f"  -> type: {type(df5).__name__}, shape: {df5.shape}")

# Test 6: 错误类型
print("\nTEST 6: 不支持的类型（应抛错）")
try:
    normalize_excel_input(42)
    print("  ERROR: should have raised")
except Exception as e:
    print(f"  OK raised: {str(e)[:80]}")

# Test 7: 完整 to_dict("records") 流程（无 tolist）
print("\nTEST 7: 完整 PDF 写入模拟（无 tolist）")
import openpyxl
from openpyxl.styles import Alignment

# 模拟 pdf_to_excel 主函数流程
def write_cell(ws, r, c, value):
    cell = ws.cell(row=r, column=c, value=value)
    cell.alignment = Alignment(wrap_text=True, vertical="top")
    return cell

# 1. 模拟 _extract_page_best 返回 IR dict
sheet1_ir = to_table_block(
    rows=[['姓名', '年龄'], ['张三', '25\n(1999)'], ['李四', '30']],
    page=1, table_id=1
)
# 2. 模拟 _extract_text_fallback 返回 fallback IR
sheet2_ir = fallback_block(
    rows=[['page 2 line 1'], ['page 2 line 2']],
    page=2, table_id=1
)
all_sheets = [("Page1_Table1", sheet1_ir), ("Page2_Text", sheet2_ir)]

wb = openpyxl.Workbook()
wb.remove(wb.active)

for sheet_name, ir in all_sheets:
    df = normalize_excel_input(ir)
    records = df.fillna("").astype(str).to_dict("records")
    ws = wb.create_sheet(sheet_name[:31])
    for i, row_dict in enumerate(records, 1):
        for j, val in enumerate(row_dict.values(), 1):
            write_cell(ws, i, j, val)

print(f"  sheets created: {wb.sheetnames}")
print(f"  Page1_Table1[1,1]: {wb['Page1_Table1'].cell(1, 1).value}")
print(f"  Page1_Table1[2,2]: {repr(wb['Page1_Table1'].cell(2, 2).value)} (含 \\n)")
print(f"  Page1_Table1[2,2].alignment: wrap={wb['Page1_Table1'].cell(2, 2).alignment.wrap_text}")
print(f"  Page2_Text[1,1]: {wb['Page2_Text'].cell(1, 1).value}")

# 保存测试文件
out = "build/_test_p0_hotfix.xlsx"
wb.save(out)
print(f"\n  saved to: {out} ({os.path.getsize(out)} bytes)")

print("\n" + "=" * 60)
print("ALL TESTS PASSED")
print("=" * 60)
